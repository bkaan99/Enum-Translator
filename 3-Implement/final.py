import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\kaan.gurgen\Desktop\Son\enum_core\Enum_Core DONE_Cleared_Final.xlsx")
sheet = wb["Sheet1"]

for i in range(1, sheet.max_row + 1):
    col2_value = sheet.cell(row=i, column=2).value
    if col2_value:
        start_index = col2_value.find('"') + 1
        end_index = col2_value.find('"', start_index)
        if end_index > start_index:
            sheet.cell(row=i, column=2).value = col2_value[:start_index] + col2_value[end_index:]

wb.save("example10.xlsx")

wb2 = openpyxl.load_workbook("example10.xlsx")
sheet = wb2["Sheet1"]

for i in range(1, sheet.max_row + 1):
    col2_value = sheet.cell(row=i, column=2).value 
    col4_value = sheet.cell(row=i, column=4).value
    if col2_value:
        if(col2_value.find('""') != -1):
            print(col4_value)
            col2_value=col2_value.replace('""', '"'+col4_value+'"')
            print(col2_value+"\n")          
    
            sheet.cell(row=i, column=2).value = col2_value
print("\n----------> Üstteki ingilizce String İfade Alttaki '""' arasındaki Türkçe ifadenin olduğu yere yazıldı.<---------- \n")
wb2.save("Enum_Core DONE_replaced.xlsx")
